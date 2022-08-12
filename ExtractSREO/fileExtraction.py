# Authors: Ryan Dunn and Cameron Gonzalez
from io import IncrementalNewlineDecoder
import os
import sys
import pandas as pd
pd.options.mode.chained_assignment = None
from prepareData import *
from trainModel import *
import warnings
import json as json
from docx import Document
import openpyxl as pyxl
import xlrd
import csv
import numpy as np
import shutil

#PDF extraction imports
import logging
from zipfile import ZipFile
from adobe.pdfservices.operation.pdfops.options.extractpdf.extract_renditions_element_type import ExtractRenditionsElementType
from adobe.pdfservices.operation.auth.credentials import Credentials
from adobe.pdfservices.operation.exception.exceptions import ServiceApiException, ServiceUsageException, SdkException
from adobe.pdfservices.operation.pdfops.options.extractpdf.extract_pdf_options import ExtractPDFOptions
from adobe.pdfservices.operation.pdfops.options.extractpdf.extract_element_type import ExtractElementType
from adobe.pdfservices.operation.execution_context import ExecutionContext
from adobe.pdfservices.operation.io.file_ref import FileRef
from adobe.pdfservices.operation.pdfops.extract_pdf_operation import ExtractPDFOperation
from adobe.pdfservices.operation.pdfops.options.extractpdf.table_structure_type import TableStructureType


logging.basicConfig(level=os.environ.get("LOGLEVEL", "INFO"))

warnings.simplefilter(action='ignore', category=UserWarning)

f = open('columnHeaderData.json')
columnHeaderData = json.load(f)
ROW, COLUMN = 0, 1 # Values Indicating DataFrame Axis
PERMITTED_FORMATS = ["csv", "xlsx", "xls", "docx", "pdf"]
HEADER_MODEL = 'Model/headerTest.pt'
modelFilePath = None

os.chdir(os.path.dirname(os.path.abspath(__file__)))


def setModelFilePath(name):
	global modelFilePath
	modelFilePath = name
def getModelFilePath():
	return modelFilePath

#Extracts an SREO file into a pandas dataframe
def extractSREO(curFilePath): 
	splitPath = curFilePath.split(".")
	fileType = splitPath[len(splitPath) - 1]
	sreoData = []
	if fileType not in PERMITTED_FORMATS:
		raise TypeError("Error: Please input compatible file type!")
	elif fileType == "csv":
		sreoData = [pd.read_csv(curFilePath, header=None)]
	#Converts xls pages to csv which can be read by pandas
	elif fileType == "xls":
		with xlrd.open_workbook(curFilePath) as workbook:
			for worksheet in workbook.sheets():
				with open('workbook.csv', 'w', newline="") as csv_file:
					col = csv.writer(csv_file)
					for row in range(worksheet.nrows):
						col.writerow(worksheet.row_values(row))
				df = pd.read_csv("workbook.csv", header=None)
				sreoData.append(df)
				os.remove("workbook.csv")
	#Import xlsx file as openpyxl workbook and fills in merged rows before converting to pandas dataframe
	elif fileType == "xlsx":
		workbook = pyxl.load_workbook(curFilePath,data_only=True)
		for worksheet in workbook.worksheets:
			for mergedCells in list(worksheet.merged_cells.ranges):
				min_col, min_row, max_col, max_row = pyxl.utils.range_boundaries(str(mergedCells))
				top_left_cell_value = worksheet.cell(row=min_row, column=min_col).value
				worksheet.unmerge_cells(str(mergedCells))
				for row in worksheet.iter_rows(min_col=min_col, min_row=min_row, max_col=min_col, max_row=max_row):
					for cell in row:
						cell.value = top_left_cell_value
			df = pd.DataFrame(worksheet.values)
			df = df.fillna(value=np.nan)
			sreoData.append(df)
	elif fileType == "docx":
		document = Document(curFilePath)
		for table in document.tables:
			df = [['' for i in range(len(table.columns))] for j in range(len(table.rows))]
			for i, row in enumerate(table.rows):
				for j, cell in enumerate(row.cells):
					if cell.text:
						df[i][j] = cell.text
			df = pd.DataFrame(df)
			sreoData.append(df)
	elif fileType == "pdf":
		sreoData = extractFromPDF(curFilePath)
	# Removes Uneccessary Information from DataFrame and Formats Correctly
	i = 0
	while i < len(sreoData):
		df = sreoData[i]
		df.replace(to_replace='/n', value=' ', regex=True, inplace=True)
		df.replace(to_replace='  ', value=' ', regex=True, inplace=True)
		df.mask(df == '', inplace=True)
		df.dropna(axis=ROW, how='all', inplace=True)
		df.dropna(axis=COLUMN, how='all', inplace=True)
		df = df.reset_index(drop=True).rename_axis(None, axis=COLUMN)
		df = transposeVerticalRows(df)
		# Reformat DataFrame to Apply Header
		index = getHeaderIndex(df)
		if index == -1:
			del sreoData[i]
		else:
			df = mergeHeaderRows(df, index)
			df.columns = df.iloc[index]
			df = df[(index + 1):].reset_index(drop=True).rename_axis(None, axis=COLUMN)
			sreoData[i] = df
			i += 1
	return sreoData

#Transposes the dataframe if each SREO is stored horizontally
def transposeVerticalRows(df):
	transposedDF = df.T
	transposedHeaderIndex = getHeaderIndex(transposedDF)
	dfHeaderIndex = getHeaderIndex(df)
	if transposedHeaderIndex == -1:
		return df
	if dfHeaderIndex == -1:
		return transposedDF
	dfCount = 0
	transposedCount = 0
	for cell in df.iloc[dfHeaderIndex]:
		if isValidColumnHeader(str(cell)):
			dfCount += 1
	for cell in transposedDF.iloc[transposedHeaderIndex]:
		if isValidColumnHeader(str(cell)):
			transposedCount += 1
	if transposedCount > dfCount:
		return transposedDF
	else:
		return df

#Checks if the header is contained on multiple rows. If it is, merges the rows
def mergeHeaderRows(df, headerIndex):
	row1 = df.iloc[headerIndex]
	row2 = df.iloc[headerIndex + 1]
	numCells = 0
	numCellsContainingHeaders = 0
	for cell in row2:
		if str(cell) != "nan":
			numCells += 1
		for headerList in columnHeaderData["headers"]:
			for header in headerList:
				if str(cell) in header and len(str(cell)) > 1:
					numCellsContainingHeaders += 1
					continue
	if(float(numCellsContainingHeaders) / float(numCells)) > 0.8:
		for i in range(len(row1)):
			if(i in row2 and i < len(row2)):
				if str(row2[i]) != "nan" :
					if str(row1[i]) == "nan":
						row1[i] = str(row2[i])
					elif str(row1[i])[len(str(row1[i]))-1] != " " or str(row2[i])[0] != " ":
						row1[i] = str(row1[i]) + " " + str(row2[i])
					else:
						row1[i] = str(row1[i]) + str(row2[i])
		df.iloc[headerIndex] = row1
		return df.drop(index=headerIndex+1)
	else:
		return df

#Uses the header row classification model HEADER_MODEL to find the index of the header row
def getHeaderIndex(df):
	for row in range(df.shape[ROW]):
		rowString = df.iloc[row].apply(str).str.cat(sep=' ')
		confidence = outputConfidence(False,HEADER_MODEL,rowString)
		if confidence[0] == "Valid":
			return row
	return -1


def fillTemplate(sreoData):
	labels = []
	highestConfidenceColumns = {}
	for headerList in columnHeaderData["headers"]:
		labels.append(headerList[0])
		highestConfidenceColumns[headerList[0]] = 0
	labels.append("Loan Type")

	templates = []
	for df in sreoData:
		sreoTemplate = pd.DataFrame(columns=labels)
		for dataColumn in df.columns:
			if type(modelFilePath) is str:
				output = outputConfidence(True, modelFilePath, str(dataColumn))
			else:
				output = outputConfidenceMultipleModels(True, modelFilePath, str(dataColumn))
			outputLabel = output[0]
			outputConfidenceVal = output[1]
			if outputLabel != "N/A" and outputConfidenceVal > highestConfidenceColumns[outputLabel]:
				highestConfidenceColumns[outputLabel] = outputConfidenceVal
				if(type(df[dataColumn]) is pd.DataFrame):
					if(len(df[dataColumn].columns) > 1):
						dataFrameColumn = df[dataColumn].iloc[:,0]	
				else:
					dataFrameColumn = df[dataColumn]
				sreoTemplate[outputLabel] = dataFrameColumn

		for row in range(df.shape[ROW]):
			for column in df.columns:
				if "Fannie Mae" in str(df.at[row, column]) or "fannie mae" in str(df.at[row, column]):
					sreoTemplate.at[row,'Loan Type'] = "Fannie Mae"
				elif "Freddie Mac" in str(df.at[row, column]) or "freddie mac" in str(df.at[row, column]):
					sreoTemplate.at[row,'Loan Type'] = "Freddie Mac"
		templates.append(sreoTemplate)
	
	return templates

#Note: Adobe Extract API is not very reliable, should work to improve it
def extractFromPDF(filePath):
	logging.basicConfig(level=os.environ.get("LOGLEVEL", "INFO"))
	try:
		for f in os.listdir("pdfExtraction/"):
			if len(f.split(".")) < 2:
				shutil.rmtree(os.path.join("pdfExtraction/", f))
			else:
				os.remove(os.path.join("pdfExtraction/", f))
		
		credentials = Credentials.service_account_credentials_builder().from_file("Extract API Credentials/pdfservices-api-credentials.json").build()
		
		execution_context = ExecutionContext.create(credentials)
		extract_pdf_operation = ExtractPDFOperation.create_new()
		
		source = FileRef.create_from_local_file(filePath)
		extract_pdf_operation.set_input(source)
		
		extract_pdf_options: ExtractPDFOptions = ExtractPDFOptions.builder().with_elements_to_extract([ExtractElementType.TEXT, ExtractElementType.TABLES]) \
		.with_element_to_extract_renditions(ExtractRenditionsElementType.TABLES) \
		.with_table_structure_format(TableStructureType.CSV) \
		.build()
		extract_pdf_operation.set_options(extract_pdf_options)
		
		result: FileRef = extract_pdf_operation.execute(execution_context)
		
		zipFilePath = "pdfExtraction/extractedPDF.zip"
		result.save_as(zipFilePath)
		with ZipFile(zipFilePath, 'r') as zipObject:
			zipObject.extractall("pdfExtraction")

		sreoData = []
		for fileName in os.listdir("pdfExtraction/tables"):
			if fileName.split(".")[len(fileName.split("."))-1] == "csv":
				sreoData.append(pd.read_csv("pdfExtraction/tables/" + fileName,header=None))
		
		return sreoData


	except (ServiceApiException, ServiceUsageException, SdkException):
		logging.exception("Exception encountered while executing operation")

def outputConfidenceMultipleModels(testColumn, modelFilePath, textInput):
	outputCounts = {}
	for model in modelFilePath:
		output = outputConfidence(testColumn, model, textInput)
		if output[0] in outputCounts.keys():
			outputCounts[output[0]] = outputCounts[output[0]] + 1
		else:
			outputCounts[output[0]] = 1

	maxCount = 0
	highestOutput = ""
	for key in outputCounts.keys():
		if outputCounts[key] > maxCount:
			maxCount = outputCounts[key]
			highestOutput = key
	if float(maxCount) / float(len(outputCounts)) >= 0.5:
		return highestOutput, float(maxCount) / float(len(outputCounts))
	else:
		return "N/A", float(maxCount) / float(len(outputCounts))

#Takes in a file path and returns a filled template of relevant data points
def standardizeSREO(sreoFilePath):
	return fillTemplate(extractSREO(sreoFilePath))

def autoCreateModel(goal):
	trainingData = pd.DataFrame(columns=['label','text'])
	count = 1
	errors = sys.maxsize
	while errors > goal:
		print("----------------------------------------------------------------------------------")
		trainingData = pd.DataFrame(columns=['label','text'])
		numReps = random.randint(35, 65) * 100
		LR = float(random.randint(35,45)) / 10.0
		print("Number of Repeats: " + str(numReps) + " Learning Rate: " + str(LR))
		modelFilePath = "newTrial" + str(count) + "-" + str(numReps) + ".pt"
		setModelFilePath(modelFilePath)
		trainingData = pd.concat([trainingData, generateData(numReps)],ignore_index=True)
		trainingData.to_csv("trainingData.csv",index=False)
		trainModel(True,True,modelFilePath,"trainingData.csv", LR)
		count += 1
		errors = testOnSolvedCSV(goal)
		if errors > goal:
			os.remove(modelFilePath)
			os.remove(modelFilePath.replace(".pt", "Vocab.pt"))
	print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
	print("New Best Model = " + modelFilePath)

def trainAgainstSolvedCSV(isNewModel, modelFilePath):
	trainingData = pd.DataFrame(columns=['label','text'])
	if isNewModel:
		for file in os.listdir("solvedCSVs/"):
			testFrame = extractSREO("solvedCSVs/" + file)[0]
			answerRow = pd.read_csv("solvedCSVs/" + file, header=0).drop(index=0).reset_index(drop=True).rename_axis(None, axis=COLUMN).replace(to_replace='\n', value=' ', regex=True)
			tempTrainingData = pd.DataFrame(columns=['label','text'])

			for i in range(len(testFrame.columns)):
				if str(testFrame.columns[i]) != "nan":
					tempTrainingData.at[i,'label'] = str(answerRow.columns[i])
					tempTrainingData.at[i,'text'] = str(testFrame.columns[i])
					if "." in tempTrainingData.at[i,'label']:
						tempTrainingData.at[i,'label'] = str(tempTrainingData.at[i,'label']).split(".")[0]
			tempTrainingData.dropna(axis=ROW, how='all', inplace=True)
			tempTrainingData.dropna(axis=COLUMN, how='all', inplace=True)
			trainingData = pd.concat([trainingData, tempTrainingData], ignore_index=True)
		trainingData.to_csv("trainingData.csv", index=False)
	else:
		trainingData = pd.read_csv("trainingData.csv")
		trainingData = trainingData.sample(frac=1)
		trainingData.to_csv("trainingData.csv",index=False)
	trainModel(True, isNewModel, modelFilePath, "trainingData.csv", 1)

################################## Testing Functions ##########################################
#Takes in a dataframe and prints out the results of inputting each column header into the model
def testConfidence(trainColumn, df):
	global modelFilePath
	highestOutputs = {}
	for column in df.columns:
		if type(modelFilePath) is str:
			output = outputConfidence(trainColumn, modelFilePath, str(column))
		else: 
			output = outputConfidenceMultipleModels(True, modelFilePath, str(column))

		outputLabel = output[0]
		outputConfidenceVal = output[1]

		if outputLabel in highestOutputs:
			if highestOutputs[outputLabel][1] < outputConfidenceVal:
				highestOutputs[outputLabel] = column, outputConfidenceVal
		elif outputLabel != "N/A":
			highestOutputs[outputLabel] = column, outputConfidenceVal

		print(str(column) + ' --> ' + outputLabel + ' ' + str(outputConfidenceVal))

	print("----------------- Highest Values --------------------")
	for output in highestOutputs.keys():
		print(highestOutputs[output][0] + " --> " + output + " " + str(highestOutputs[output][1]))


def testOnSolvedCSV(goal=sys.maxsize):
	validLabels = []
	startTime = time.perf_counter()
	totalCorrect, totalNum = 0, 0
	missingCounts, incorrectMatchCounts = {}, {}
	trainedToRecognizeButNotUsed = ["Loan Type", "Loan Amount", "Spread", "Index"] #Loan Type is used, is not extracted by the model so it is excluded here
	for headerList in columnHeaderData["headers"]:
		if headerList[0] not in trainedToRecognizeButNotUsed:
			validLabels.append(headerList[0])

	for file in os.listdir("solvedCSVs/"):
		CSVDataFrame = pd.read_csv("solvedCSVs/" + file,header=None)
		CSVDataFrame.replace(to_replace='/n', value=' ', regex=True, inplace=True)
		CSVDataFrame.replace(to_replace='  ', value=' ', regex=True, inplace=True)
		
		correct, numVals = 0, 0

		missing, incorrectMatch = [], []
		highestClassifications = {}
		for columnIndex in range(CSVDataFrame.shape[COLUMN]):
			columnLabel = str(CSVDataFrame.at[1, columnIndex])
			output, confidence = outputConfidence(True, modelFilePath, columnLabel)
			if output in highestClassifications:
				if confidence > highestClassifications[output][1]:
					highestClassifications[output] = columnLabel, confidence
			else:
				highestClassifications[output] = columnLabel, confidence

		for columnIndex in range(CSVDataFrame.shape[COLUMN]):
			correctLabel = str(CSVDataFrame.at[0,columnIndex])
			if correctLabel not in trainedToRecognizeButNotUsed:
				actualLabel = str(CSVDataFrame.at[1,columnIndex])
				if correctLabel != "nan" and correctLabel != "N/A":
					if correctLabel not in highestClassifications:
						missing.append(correctLabel)
						if correctLabel in missingCounts:
							missingCounts[correctLabel] += 1
						else:
							missingCounts[correctLabel] = 1
					elif highestClassifications[correctLabel][0] == actualLabel:
						correct += 1
						totalCorrect += 1
					else:
						incorrectMatch.append(correctLabel)
						if correctLabel in incorrectMatchCounts:
							missingCounts[correctLabel] += 1
						else:
							missingCounts[correctLabel] = 1
					totalNum += 1
					numVals += 1
		print(file)
		print("Required Headers Accuracy --> " + str("{:.2%}".format(correct/numVals)))
		print()
		if len(missing) != 0:
			print("Missing: ")
			for missed in missing:
				print(missed)
			print()
		if len(incorrectMatch) != 0:
			print("Incorrect Match: ")
			for missed in incorrectMatch:
				print(missed)
			print()
		print("----------------------------------------------------------------------------------------")
		if totalNum - totalCorrect > goal:
			print("Greater errors than goal, generating new model")
			return totalNum - totalCorrect

	print("----------------------------------------------------------------------------------------")
	totalErrors = 0
	print("Headers Missing:")
	for missed in missingCounts:
		print(missed + " = " + str(missingCounts[missed]))
		totalErrors += missingCounts[missed]
	print()
	print("Incorrect Matches:")
	for missed in incorrectMatchCounts:
		print(missed + " = " + str(incorrectMatch[missed]))
		totalErrors += incorrectMatchCounts[missed]

	print()
	print("Total Required Headers Accuracy --> " + str("{:.2%}".format(totalCorrect/totalNum)))
	print("Total Errors --> " + str(totalErrors))
	print("Total Processing Time: " + str(time.perf_counter() - startTime) + "s")
	print("----------------------------------------------------------------------------------------")

	return totalErrors
	

if __name__ == "__main__":
	running = ""
	setModelFilePath(input("Current model file path: "))
	while running != "0":
		running = input("1 for file extraction, 2 for test on solved csv, 3 for data generation, 4 for training, 5 for more, 0 to quit: ")
		if running == "5":
			running = input("5 to auto-generate model, 6 to train on solved csv, 7 to change model: ")
		if running == "1":
			filePath = input("Input file path: ")
			sreoData = standardizeSREO(filePath)
			exportPath = input("Export file name: ")
			if len(sreoData) == 1:
				sreoData[0].to_csv(exportPath + ".csv",index=False)
				sreoData[0].to_json(exportPath + ".json")
			else:
				count = 0
				for df in sreoData:
					df.to_csv(str(count) + "-" + exportPath + ".csv", index=False)
					df.to_json(str(count) + "-" + exportPath + ".json")
					count += 1
		elif running == "2":
				testOnSolvedCSV()
		elif running == "3":
			numRepeats = int(input("Enter number of repeats: "))
			trainColumn = input("1 for column model 2 for header model: ") == "1"
			if trainColumn:
				generateData(numRepeats).to_csv("trainingData.csv")
			else:
				generateHeaderData(numRepeats).to_csv("trainingData.csv")
		elif running == "4":
			modelName = input("Model File Path: ")
			trainColumn = input("1 for column model 2 for header model: ") == "1"
			newModel = input("1 for new model 2 for existing model: ") == "1"
			trainModel(trainColumn, newModel, modelName, "trainingData.csv")
		elif running == "5":
			goal = int(input("Errors to beat: "))
			autoCreateModel(goal)
		elif running == "6":
			newModel = input("1 for new model, 2 for existing model")
			filePath = input("Model file path: ")
			trainAgainstSolvedCSV(newModel, filePath)
		elif running == "7":
			setModelFilePath(input("Input File Path: "))
		print()